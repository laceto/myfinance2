#!/usr/bin/env python3
"""
get_insights.py

Daily trading insights for Italian equities.

Input:  data/results/it/analysis_results.parquet
Joins:  sectors.xlsx       (symbol → name, sector)
        marginabili.xlsx   (name → marginabile flag)
Output: signals/daily_brief.xlsx   — single workbook, one sheet per view
        signals/daily_brief.txt    — console-ready morning brief (also printed to stdout)

Scoring
-------
SIGNAL_COLS (10 method columns) are each ∈ {-1, 0, 1}.
  score       = sum of SIGNAL_COLS for the last bar   → range [−10, +10]
  score_delta = score_today − score_yesterday         → positive = strengthening
  trend_bonus = 1 if OLS(close ~ time_id) is significant and directionally aligned
  conviction  = score + score_delta × 0.5 + trend_bonus − staleness_penalty

rrg is used as the bull/bear regime gate (not counted in score).

Usage
-----
    python get_insights.py
    python get_insights.py --no-lm                    # skip OLS, fast mode
    python get_insights.py --min-score 5 --top-n 15
    python get_insights.py --max-days 20              # max days since last signal change
"""

from __future__ import annotations

import argparse
import glob
import logging
import sys
import warnings

# Windows cp1252 consoles cannot encode UTF-8 characters used in log/print output.
sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")
from datetime import date
from pathlib import Path
from typing import List

import numpy as np
import pandas as pd
import statsmodels.api as sm
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# CLI arguments
# ---------------------------------------------------------------------------

parser = argparse.ArgumentParser(description="Daily trading insights for Italian equities.")
parser.add_argument("--no-lm",     action="store_true", help="Skip OLS linear trend models (faster).")
parser.add_argument("--min-score", type=int, default=3,  help="Minimum signal score to include in brief (default 3).")
parser.add_argument("--top-n",     type=int, default=30, help="Max candidates per section in brief (default 30).")
parser.add_argument("--max-days",  type=int, default=60, help="Max days since last signal change (default 60).")
args = parser.parse_args()

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    stream=sys.stdout,
)
log = logging.getLogger(__name__)
warnings.filterwarnings("ignore", category=RuntimeWarning, module="statsmodels")

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

# 10 trading-method signal columns — each ∈ {-1, 0, 1}, max score = 10.
# rrg is kept separate as the bull/bear regime gate.
SIGNAL_COLS: List[str] = [
    "rtt_5020",
    "rsma_50100",
    "rsma_100150",
    "rsma_50100150",
    "rema_50100",
    "rema_100150",
    "rema_50100150",
    "rbo_20",
    "rbo_50",
    "rbo_150",
]

# Columns to detect regime changes in (includes rrg flip events)
CHANGE_COLS: List[str] = SIGNAL_COLS + ["rrg"]

LM_WINDOWS: List[int] = [7, 15, 30, 45]

PARQUET_PATH  = Path("data/results/it/analysis_results.parquet")
EXCEL_PATH    = Path("signals/daily_brief.xlsx")
BRIEF_PATH    = Path("signals/daily_brief.txt")
SIGNALS_DIR   = Path("signals")

SIGNALS_DIR.mkdir(exist_ok=True)

# ---------------------------------------------------------------------------
# Pure helpers
# ---------------------------------------------------------------------------


def semi_join(left: pd.DataFrame, right: pd.DataFrame, on: str | List[str]) -> pd.DataFrame:
    """Return rows of *left* that have a matching key in *right* (R dplyr::semi_join)."""
    keys = [on] if isinstance(on, str) else on
    return left.merge(right[keys].drop_duplicates(), on=keys, how="inner")


def get_last_swing(df: pd.DataFrame, swing: str) -> pd.DataFrame:
    """Per symbol, return the last row where *swing* is not NA."""
    return (
        df.dropna(subset=[swing])
        .groupby("symbol", group_keys=False)
        .last()
        .reset_index()
        .sort_values("date", ascending=False)
        [["symbol", "name", "date", swing]]
        .rename(columns={"date": "date_last_swing"})
    )


def detect_change(df: pd.DataFrame, regime: str) -> pd.DataFrame:
    """Within a single-symbol group, return the last row where *regime* value changed."""
    cols = [c for c in ["name", "symbol", "sector", "marginabile", "date", "volume", regime] if c in df.columns]
    result = df[cols].copy()
    result["change"] = (result[regime] != result[regime].shift(1)).astype(int)
    return result[result["change"] == 1].tail(1)


def compute_changes(df: pd.DataFrame, symbol_keys: pd.DataFrame) -> pd.DataFrame:
    """
    For every column in CHANGE_COLS, detect the last regime-change row per symbol.

    Uses explicit groupby iteration — pandas 3.0 excludes the groupby key from the
    group DataFrame inside apply(), so we re-inject it manually.
    """
    filtered = semi_join(df, symbol_keys, on="symbol")
    frames: List[pd.DataFrame] = []
    for col in CHANGE_COLS:
        if col not in filtered.columns:
            continue
        col_frames: List[pd.DataFrame] = []
        for sym, group in filtered.groupby("symbol"):
            g = group.copy()
            g["symbol"] = sym
            row = detect_change(g, col)
            if not row.empty:
                col_frames.append(row)
        if col_frames:
            frames.append(pd.concat(col_frames, ignore_index=True).sort_values("date", ascending=False))
    if not frames:
        return pd.DataFrame()
    return pd.concat(frames, ignore_index=True).sort_values("date", ascending=False)


def compute_score(df: pd.DataFrame, symbol_keys: pd.DataFrame) -> pd.DataFrame:
    """
    For the last bar of each symbol: sum SIGNAL_COLS into a score, capture volume.
    Score range: [−len(SIGNAL_COLS), +len(SIGNAL_COLS)].
    """
    last_bar = semi_join(df, symbol_keys, on="symbol").groupby("symbol", group_keys=False).tail(1)
    signal_cols_present = [c for c in SIGNAL_COLS if c in last_bar.columns]
    long = last_bar.melt(
        id_vars=["symbol", "sector", "name", "marginabile", "volume"],
        value_vars=signal_cols_present,
        var_name="method",
        value_name="signal",
    )
    return (
        long
        .groupby(["symbol", "sector", "name", "marginabile"], as_index=False)
        .agg(score=("signal", "sum"), last_day_volume=("volume", "mean"))
        .sort_values(["score", "last_day_volume"], ascending=[False, False])
        .reset_index(drop=True)
    )


def compute_score_delta(df: pd.DataFrame, symbol_keys: pd.DataFrame) -> pd.DataFrame:
    """
    Return per-symbol score delta (score_today − score_yesterday).
    Positive delta means more methods flipped bullish/bearish today.
    """
    signal_cols_present = [c for c in SIGNAL_COLS if c in df.columns]
    daily = (
        semi_join(df, symbol_keys, on="symbol")
        .melt(
            id_vars=["symbol", "date"],
            value_vars=signal_cols_present,
            var_name="method",
            value_name="signal",
        )
        .groupby(["symbol", "date"], as_index=False)
        .agg(score=("signal", lambda s: s.sum(skipna=True)))
        .sort_values(["symbol", "date"], ascending=[True, False])
    )
    result = []
    for sym, grp in daily.groupby("symbol"):
        grp = grp.reset_index(drop=True)
        delta = int(grp["score"].iloc[0] - grp["score"].iloc[1]) if len(grp) >= 2 else 0
        result.append({"symbol": sym, "score_delta": delta})
    return pd.DataFrame(result)


def compute_score_ts(df: pd.DataFrame, symbol_keys: pd.DataFrame) -> pd.DataFrame:
    """Daily aggregate score per symbol (marginabile only), long format."""
    signal_cols_present = [c for c in SIGNAL_COLS if c in df.columns]
    return (
        semi_join(df, symbol_keys, on="symbol")
        .melt(
            id_vars=["symbol", "name", "marginabile", "date"],
            value_vars=signal_cols_present,
            var_name="method",
            value_name="signal",
        )
        .groupby(["symbol", "name", "marginabile", "date"], as_index=False)
        .agg(score=("signal", lambda s: s.sum(skipna=True)))
        .sort_values(["symbol", "name", "date"], ascending=[True, True, False])
        .query("marginabile == 'si'")
        .drop(columns=["marginabile"])
        .reset_index(drop=True)
    )


def linear_model(df: pd.DataFrame) -> pd.DataFrame:
    """OLS(close ~ time_id) on the last *window_lm* rows → slope, p-value, adj-R²."""
    window_lm = int(df["window"].iloc[0])
    data = df.tail(window_lm).copy()
    data["time_id"] = range(1, len(data) + 1)
    try:
        model = sm.OLS(data["close"], sm.add_constant(data["time_id"])).fit()
        estimate      = model.params["time_id"]
        p_value       = model.pvalues["time_id"]
        adj_r_squared = model.rsquared_adj
    except Exception:
        estimate = p_value = adj_r_squared = np.nan
    return pd.DataFrame([{
        "rrg":             data["rrg"].iloc[0] if not data.empty else np.nan,
        "symbol":          data["symbol"].iloc[0],
        "window_lm":       window_lm,
        "marginabile":     data["marginabile"].iloc[0] if not data.empty else np.nan,
        "name":            data["name"].iloc[0],
        "min_date_window": data["date"].min(),
        "estimate":        estimate,
        "p_value":         p_value,
        "adj_r2":          adj_r_squared,
    }])


def build_conviction(
    df: pd.DataFrame,
    score_col: str,
    delta_df: pd.DataFrame,
    trend_df: pd.DataFrame | None,
    today: pd.Timestamp,
    direction: str,  # "bull" or "bear"
) -> pd.DataFrame:
    """
    Attach score_delta, trend_bonus, days_since_change, and conviction to *df*.

    conviction = score
                 + score_delta × 0.5       (momentum bonus)
                 + trend_bonus × 1.0       (OLS-confirmed trend)
                 − staleness × 0.5         (days_since_change / 30, capped at 2)

    Invariant: conviction is always NaN-safe and rounded to 2dp.
    """
    result = df.copy()

    # Score delta
    result = result.merge(delta_df, on="symbol", how="left")
    result["score_delta"] = result["score_delta"].fillna(0)

    # Trend bonus
    if trend_df is not None and not trend_df.empty:
        result = result.merge(trend_df[["symbol", "adj_r2"]].rename(columns={"adj_r2": "best_adj_r2"}),
                              on="symbol", how="left")
        result["trend_bonus"] = result["best_adj_r2"].notna().astype(float)
    else:
        result["best_adj_r2"] = np.nan
        result["trend_bonus"] = 0.0

    # Staleness
    if "date_last_change" in result.columns:
        result["days_since_change"] = (today - pd.to_datetime(result["date_last_change"])).dt.days
    else:
        result["days_since_change"] = np.nan
    staleness = (result["days_since_change"].fillna(60) / 30).clip(0, 2)

    # Conviction
    result["conviction"] = (
        result[score_col].fillna(0)
        + result["score_delta"] * 0.5
        + result["trend_bonus"]
        - staleness * 0.5
    ).round(2)

    return result.sort_values("conviction", ascending=(direction == "bear"))


# ---------------------------------------------------------------------------
# Excel formatting helpers
# ---------------------------------------------------------------------------

HEADER_FILL  = PatternFill(fill_type="solid", fgColor="1F4E79")
HEADER_FONT  = Font(bold=True, color="FFFFFF", size=10)
BULL_FILL    = PatternFill(fill_type="solid", fgColor="E8F5E9")
BEAR_FILL    = PatternFill(fill_type="solid", fgColor="FFEBEE")
NEUTRAL_FILL = PatternFill(fill_type="solid", fgColor="FFF9C4")
BOLD_FONT    = Font(bold=True, size=10)


def _autofit_columns(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(cell.value)) if cell.value is not None else 0) for cell in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 40)


def _style_sheet(ws, fill: PatternFill) -> None:
    """Apply header style and alternating row fill."""
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        if row_idx % 2 == 0:
            for cell in row:
                cell.fill = fill
    ws.freeze_panes = "A2"
    _autofit_columns(ws)


def save_excel(sheets: dict[str, pd.DataFrame], path: Path) -> None:
    """Write multiple DataFrames to a single .xlsx workbook, one sheet each."""
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        for sheet_name, df in sheets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    # Post-process with openpyxl for styling
    wb = load_workbook(path)
    fills = {
        "bull_candidates":  BULL_FILL,
        "bear_candidates":  BEAR_FILL,
        "regime_changes":   NEUTRAL_FILL,
        "linear_bull":      BULL_FILL,
        "linear_bear":      BEAR_FILL,
        "max_equity":       NEUTRAL_FILL,
        "volume_movers":    NEUTRAL_FILL,
        "bull_ts_score":    BULL_FILL,
        "bear_ts_score":    BEAR_FILL,
    }
    for ws in wb.worksheets:
        _style_sheet(ws, fills.get(ws.title, NEUTRAL_FILL))
    wb.save(path)
    log.info("Excel workbook saved -> %s  (%d sheets)", path, len(sheets))


# ---------------------------------------------------------------------------
# Load & enrich
# ---------------------------------------------------------------------------

log.info("Loading parquet: %s", PARQUET_PATH)
if not PARQUET_PATH.exists():
    sys.exit(f"Error: {PARQUET_PATH} not found.")
output_signal = pd.read_parquet(PARQUET_PATH)
output_signal["date"] = pd.to_datetime(output_signal["date"]).dt.normalize()
log.info("Loaded %d rows × %d cols, %d symbols.",
         len(output_signal), len(output_signal.columns),
         output_signal["symbol"].nunique())

# Sectors
sectors_files = glob.glob("sectors.xlsx")
if not sectors_files:
    sys.exit("Error: sectors.xlsx not found.")
sectors = pd.read_excel(sectors_files[0])
if "ticker" in sectors.columns:
    sectors = sectors.rename(columns={"ticker": "symbol"})

# Marginabili
marginabili_files = glob.glob("marginabili.xlsx")
if not marginabili_files:
    sys.exit("Error: marginabili.xlsx not found.")
marginabili = (
    pd.read_excel(marginabili_files[0])
    .groupby("Descrizione").size().reset_index(name="n")
    .sort_values("n", ascending=False)
    .assign(marginabile="si")
)

output_signal = (
    output_signal
    .merge(sectors, on="symbol", how="left")
    .merge(marginabili[["Descrizione", "marginabile"]], left_on="name", right_on="Descrizione", how="left")
    .drop(columns=["Descrizione"], errors="ignore")
)

# Verify all requested signal columns are present
missing = [c for c in SIGNAL_COLS if c not in output_signal.columns]
if missing:
    log.warning("Signal columns not found in parquet (will be skipped): %s", missing)
    SIGNAL_COLS[:] = [c for c in SIGNAL_COLS if c in output_signal.columns]
    CHANGE_COLS[:] = SIGNAL_COLS + ["rrg"]

log.info("Active signal columns (%d): %s", len(SIGNAL_COLS), SIGNAL_COLS)

# ---------------------------------------------------------------------------
# Bull / Bear sets  (last bar per symbol; rrg == 1 is bull)
# ---------------------------------------------------------------------------

_last_bar = output_signal.groupby(["symbol", "name"], group_keys=False).tail(1)
bull = _last_bar.query("rrg == 1")[["symbol"]].drop_duplicates()
bear = _last_bar.query("rrg != 1")[["symbol"]].drop_duplicates()
log.info("Bull symbols: %d   Bear symbols: %d", len(bull), len(bear))

today = output_signal["date"].max()
log.info("Latest data date: %s", today.date())

# ---------------------------------------------------------------------------
# Regime changes
# ---------------------------------------------------------------------------

log.info("Computing regime changes...")
bull_changes = compute_changes(output_signal, bull)
bear_changes = compute_changes(output_signal, bear)

regime_change = pd.DataFrame()
if not bull_changes.empty or not bear_changes.empty:
    _rc_frames = []
    if not bull_changes.empty:
        _rc_frames.append(bull_changes.query("rrg.notna()").sort_values("rrg"))
    if not bear_changes.empty:
        _rc_frames.append(bear_changes.query("rrg.notna() and marginabile.notna()").sort_values("rrg"))
    if _rc_frames:
        regime_change = (
            pd.concat(_rc_frames, ignore_index=True)
            .sort_values("date", ascending=False)
            .rename(columns={"date": "date_regime_change"})
            [["name", "symbol", "sector", "marginabile", "date_regime_change", "rrg"]]
        )

# ---------------------------------------------------------------------------
# Scores, deltas, and swing levels
# ---------------------------------------------------------------------------

log.info("Computing scores...")
bull_score    = compute_score(output_signal, bull)
bear_score    = compute_score(output_signal, bear)
bull_delta    = compute_score_delta(output_signal, bull)
bear_delta    = compute_score_delta(output_signal, bear)
bull_score_ts = compute_score_ts(output_signal, bull)
bear_score_ts = compute_score_ts(output_signal, bear)

bull_swing = get_last_swing(semi_join(output_signal, bull, on="symbol"), "rl3")
bear_swing = get_last_swing(semi_join(output_signal, bear, on="symbol"), "rh3")

# Most-recent change per symbol (for days_since_change)
def _last_change(changes_df: pd.DataFrame) -> pd.DataFrame:
    if changes_df.empty:
        return pd.DataFrame(columns=["name", "symbol", "sector", "marginabile", "date_last_change"])
    df = (
        changes_df
        .rename(columns={"date": "date_last_change"})
        .drop(columns=["volume", "change"], errors="ignore")
    )
    # Explicit iteration — pandas 3.0 excludes groupby keys from group DataFrames.
    rows: List[pd.DataFrame] = []
    for sym, group in df.groupby("symbol"):
        g = group.copy()
        g["symbol"] = sym
        rows.append(g.sort_values("date_last_change", ascending=False).head(1))
    return pd.concat(rows, ignore_index=True) if rows else pd.DataFrame()

bull_last_change = _last_change(bull_changes)
bear_last_change = _last_change(bear_changes)

# Daily returns (last bar)
def _daily_return(df: pd.DataFrame, keys: pd.DataFrame) -> pd.DataFrame:
    return (
        semi_join(df, keys, on="symbol")
        [["symbol", "name", "rrg", "marginabile", "date", "close"]]
        .assign(daily_return=lambda x: x.groupby("symbol")["close"]
                .transform(lambda s: s / s.shift(1) - 1))
        .groupby("symbol", group_keys=False).tail(1)
        .reset_index(drop=True)
    )

bull_returns = _daily_return(output_signal, bull)
bear_returns = _daily_return(output_signal, bear)

# ---------------------------------------------------------------------------
# Build bull_tot / bear_tot  (one row per symbol with all context)
# ---------------------------------------------------------------------------

def _build_tot(
    score_df: pd.DataFrame,
    swing_df: pd.DataFrame,
    swing_col: str,
    last_change_df: pd.DataFrame,
    returns_df: pd.DataFrame,
) -> pd.DataFrame:
    change_cols = ["symbol", "date_last_change"] + [
        c for c in ["name", "sector", "marginabile"] if c in last_change_df.columns
    ]
    return (
        score_df
        .merge(swing_df[["symbol", "date_last_swing", swing_col]], on="symbol", how="left")
        .merge(
            last_change_df[[c for c in change_cols if c in last_change_df.columns]],
            on=["symbol"] + [c for c in ["name", "sector", "marginabile"] if c in last_change_df.columns and c in score_df.columns],
            how="left",
        )
        .merge(returns_df[["symbol", "daily_return"]], on="symbol", how="left")
    )

bull_tot = _build_tot(bull_score, bull_swing, "rl3", bull_last_change, bull_returns)
bear_tot = _build_tot(bear_score, bear_swing, "rh3", bear_last_change, bear_returns)

# ---------------------------------------------------------------------------
# Linear trend models  (skipped with --no-lm)
# ---------------------------------------------------------------------------

lm_mod        = pd.DataFrame()
bull_trend_df = None
bear_trend_df = None

if not args.no_lm:
    log.info("Running OLS linear trend models (%d windows × %d symbols)...",
             len(LM_WINDOWS), output_signal["symbol"].nunique())
    _lm_frames: List[pd.DataFrame] = []
    for w in LM_WINDOWS:
        tmp = output_signal.copy()
        tmp["window"] = w
        _lm_frames.append(tmp)
    output_signal_lm = pd.concat(_lm_frames, ignore_index=True)

    _lm_results: List[pd.DataFrame] = []
    for (sym, win), group in output_signal_lm.groupby(["symbol", "window"]):
        g = group.copy()
        g["symbol"] = sym
        g["window"] = win
        _lm_results.append(linear_model(g))
    lm_mod = pd.concat(_lm_results, ignore_index=True) if _lm_results else pd.DataFrame()
    log.info("OLS complete: %d models fitted.", len(lm_mod))

    if not lm_mod.empty:
        bull_trend_df = (
            lm_mod.query("rrg == 1 and estimate > 0 and p_value <= 0.05")
            .groupby("symbol", as_index=False)["adj_r2"].max()
        )
        bear_trend_df = (
            lm_mod.query("rrg != 1 and estimate < 0 and p_value <= 0.10")
            .groupby("symbol", as_index=False)["adj_r2"].max()
        )
else:
    log.info("--no-lm: skipping OLS linear trend models.")

# ---------------------------------------------------------------------------
# Conviction scores
# ---------------------------------------------------------------------------

bull_candidates = build_conviction(bull_tot, "score", bull_delta, bull_trend_df, today, "bull")
bear_candidates = build_conviction(bear_tot, "score", bear_delta, bear_trend_df, today, "bear")

# Apply trader filters
bull_candidates = bull_candidates[
    (bull_candidates["score"] >= args.min_score) &
    (bull_candidates["days_since_change"].fillna(999) <= args.max_days)
].head(args.top_n).reset_index(drop=True)

bear_candidates = bear_candidates[
    (bear_candidates["marginabile"] == "si") &
    (bear_candidates["score"].abs() >= args.min_score) &
    (bear_candidates["days_since_change"].fillna(999) <= args.max_days)
].head(args.top_n).reset_index(drop=True)

# ---------------------------------------------------------------------------
# Max-equity method per symbol
# ---------------------------------------------------------------------------

_equity_cols = [c for c in output_signal.columns if c.endswith("_PL_cum") and "_fx" not in c]
max_equity   = pd.DataFrame()

if _equity_cols:
    _last_per_sym = (
        output_signal.sort_values("date", ascending=False)
        .groupby(["name", "symbol"], group_keys=False).head(1)
        [["symbol", "name", "rrg", "date"] + _equity_cols]
    )
    _eq_long = _last_per_sym.melt(
        id_vars=["symbol", "name", "rrg", "date"],
        value_vars=_equity_cols, var_name="method", value_name="equity",
    )
    _best = (
        _eq_long.sort_values(["symbol", "equity"], ascending=[True, False])
        .groupby("symbol", group_keys=False).head(1)
        .sort_values("equity", ascending=False)
        .copy()
    )
    _best["method"] = _best["method"].str.replace(r"_PL_cum$", "", regex=True)
    max_equity = _best[["symbol", "name", "rrg", "method", "equity"]].reset_index(drop=True)

# ---------------------------------------------------------------------------
# Last-day volume
# ---------------------------------------------------------------------------

last_day_volume = (
    output_signal[output_signal["date"] == today]
    .sort_values("volume", ascending=False)
    [["date", "symbol", "name", "sector", "volume"]]
    .reset_index(drop=True)
)

# ---------------------------------------------------------------------------
# Excel workbook  — single output file
# ---------------------------------------------------------------------------

# Column order for the candidate sheets
_BULL_COLS = [
    "conviction", "score", "score_delta", "trend_bonus", "days_since_change",
    "symbol", "name", "sector",
    "date_last_change", "date_last_swing", "rl3",
    "daily_return", "last_day_volume", "best_adj_r2",
]
_BEAR_COLS = [
    "conviction", "score", "score_delta", "trend_bonus", "days_since_change",
    "symbol", "name", "sector", "marginabile",
    "date_last_change", "date_last_swing", "rh3",
    "daily_return", "last_day_volume", "best_adj_r2",
]

def _safe_select(df: pd.DataFrame, cols: List[str]) -> pd.DataFrame:
    return df[[c for c in cols if c in df.columns]]

_rc_sheet = (
    regime_change[regime_change["date_regime_change"] >= today - pd.Timedelta(days=30)]
    if not regime_change.empty else pd.DataFrame()
)

_lm_bull_sheet = pd.DataFrame()
_lm_bear_sheet = pd.DataFrame()
if not lm_mod.empty:
    _lm_bull_sheet = (
        lm_mod.query("rrg == 1 and estimate > 0 and p_value <= 0.05")
        .sort_values(["window_lm", "adj_r2"], ascending=[True, False])
    )
    _lm_bear_sheet = (
        lm_mod.query("rrg != 1 and estimate < 0 and p_value <= 0.10 and marginabile == 'si'")
        .sort_values(["window_lm", "adj_r2"], ascending=[True, False])
    )

sheets = {
    "bull_candidates": _safe_select(bull_candidates, _BULL_COLS),
    "bear_candidates": _safe_select(bear_candidates, _BEAR_COLS),
    "regime_changes":  _rc_sheet,
    "linear_bull":     _lm_bull_sheet,
    "linear_bear":     _lm_bear_sheet,
    "max_equity":      max_equity,
    "volume_movers":   last_day_volume,
    "bull_ts_score":   bull_score_ts,
    "bear_ts_score":   bear_score_ts,
}

save_excel(sheets, EXCEL_PATH)

# ---------------------------------------------------------------------------
# Console morning brief
# ---------------------------------------------------------------------------

SEP_WIDE   = "=" * 100
SEP_NARROW = "-" * 60

_BRIEF_BULL_COLS = ["conviction", "score", "score_delta", "symbol", "name",
                    "sector", "days_since_change", "rl3", "daily_return"]
_BRIEF_BEAR_COLS = ["conviction", "score", "score_delta", "symbol", "name",
                    "sector", "days_since_change", "rh3", "daily_return"]

def _fmt_df(df: pd.DataFrame, cols: List[str]) -> str:
    sub = _safe_select(df, cols)
    if sub.empty:
        return "  (none)"
    # Format floats for readability
    for c in ["conviction", "score_delta", "daily_return", "rl3", "rh3"]:
        if c in sub.columns:
            sub = sub.copy()
            sub[c] = sub[c].map(lambda v: f"{v:.3f}" if pd.notna(v) else "")
    for c in ["days_since_change", "score"]:
        if c in sub.columns:
            sub = sub.copy()
            sub[c] = sub[c].map(lambda v: f"{int(v)}" if pd.notna(v) else "")
    return sub.to_string(index=False)


lines = [
    "",
    SEP_WIDE,
    f"  DAILY BRIEF — {today.date()}   "
    f"(min_score={args.min_score}  max_days={args.max_days}  top_n={args.top_n}  "
    f"{'LM=off' if args.no_lm else 'LM=on'})",
    SEP_WIDE,
    "",
    f"  BULL CANDIDATES  ({len(bull_candidates)} stocks)  "
    f"— ranked by conviction (score + momentum + trend - staleness)",
    SEP_NARROW,
    _fmt_df(bull_candidates, _BRIEF_BULL_COLS),
    "",
    f"  BEAR CANDIDATES  ({len(bear_candidates)} stocks, marginabile only)",
    SEP_NARROW,
    _fmt_df(bear_candidates, _BRIEF_BEAR_COLS),
    "",
    f"  REGIME FLIPS — last 30 days  ({len(_rc_sheet)} events)",
    SEP_NARROW,
    _rc_sheet[["date_regime_change", "symbol", "name", "rrg"]].to_string(index=False)
    if not _rc_sheet.empty else "  (none)",
    "",
    f"  VOLUME MOVERS — {today.date()}  (top 15)",
    SEP_NARROW,
    last_day_volume.head(15)[["symbol", "name", "sector", "volume"]].to_string(index=False)
    if not last_day_volume.empty else "  (none)",
    "",
    SEP_WIDE,
    f"  Full detail -> {EXCEL_PATH}",
    SEP_WIDE,
    "",
]

brief_text = "\n".join(lines)
print(brief_text)

BRIEF_PATH.write_text(brief_text, encoding="utf-8")
log.info("Morning brief saved -> %s", BRIEF_PATH)
