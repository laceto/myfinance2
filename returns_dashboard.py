"""
returns_dashboard.py — Multi-window returns dashboard for Borsa Italiana tickers.

Reads analysis_results.parquet and produces an Excel dashboard showing
per-ticker percentage returns across 8 lookback windows (3, 5, 7, 10, 15, 30,
45, 60 trading days) for both:
    - close:  absolute EUR price
    - rclose: price relative to FTSEMIB.MI benchmark

Output:
    data/results/it/returns_dashboard.xlsx

Sheets:
    Returns — one row per ticker, 16 return columns + metadata

Usage:
    python returns_dashboard.py

Invariants:
    - "today" = last available date per ticker (no lookahead bias).
    - Windows exceeding available history yield NaN (not an error).
    - Output is sorted by ticker (ascending).
"""

import logging
import sys
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from returns_calculator import LOOKBACK_WINDOWS, build_returns_dashboard

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

RESULTS_PATH = Path("data/results/it/analysis_results.parquet")
OUTPUT_PATH = Path("data/results/it/returns_dashboard.xlsx")

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------


def load_results(path: Path) -> dict[str, pd.DataFrame]:
    """
    Load analysis_results.parquet and partition by symbol.

    Returns:
        symbol -> DataFrame mapping. Each DataFrame is sorted by date ascending.

    Raises:
        FileNotFoundError: If the parquet file does not exist.
    """
    if not path.exists():
        raise FileNotFoundError(
            f"Results file not found: {path}\n"
            "Run analyze_stock.py first to generate analysis_results.parquet."
        )
    log.info("Loading analysis results from %s", path)
    df = pd.read_parquet(path)
    if "date" in df.columns:
        df["date"] = pd.to_datetime(df["date"])
    data_dict = {
        symbol: grp.sort_values("date").reset_index(drop=True)
        for symbol, grp in df.groupby("symbol")
    }
    log.info("Loaded %d symbols", len(data_dict))
    return data_dict


# ---------------------------------------------------------------------------
# Excel formatting
# ---------------------------------------------------------------------------

#: Column groups that receive a red→white→green color scale.
_RETURN_COL_PREFIXES = ("ret_close_", "ret_rclose_", "cumret_close_", "cumret_rclose_")

#: Header fill: dark-blue (same palette as trending_dashboard.py).
_HEADER_FILL = PatternFill(fill_type="solid", fgColor="1F3864")
_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_ALIGN = Alignment(horizontal="center")


def _is_return_col(col_name: str) -> bool:
    return any(col_name.startswith(p) for p in _RETURN_COL_PREFIXES)


def save_returns_dashboard(df: pd.DataFrame, output_path: Path) -> None:
    """
    Write the returns dashboard to Excel with conditional formatting.

    Formatting applied:
      - Sheet name: "Returns"
      - Header row: dark-blue background, white bold text, centred
      - Header row frozen (row 1)
      - Each return column: red(min) → white(0) → green(max) color scale
      - Column widths auto-fitted (capped at 20 chars)
      - Percentage format (0.00%) applied to all return columns

    Args:
        df:          Wide-format DataFrame from build_returns_dashboard().
        output_path: Destination .xlsx path. Parent directories are created.
    """
    output_path.parent.mkdir(parents=True, exist_ok=True)

    # --- Write raw data first ---
    df.to_excel(output_path, index=False, sheet_name="Returns")
    log.info("Raw data written to %s", output_path)

    # --- Open workbook for formatting ---
    wb = load_workbook(output_path)
    ws = wb["Returns"]
    max_row = ws.max_row

    # --- Header row ---
    for cell in ws[1]:
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = _HEADER_ALIGN

    # --- Per-column formatting ---
    for col_idx, header_cell in enumerate(ws[1], start=1):
        col_name: str = header_cell.value or ""
        col_letter = get_column_letter(col_idx)

        if _is_return_col(col_name) and max_row > 1:
            data_range = f"{col_letter}2:{col_letter}{max_row}"

            # Percentage number format
            for row_idx in range(2, max_row + 1):
                ws[f"{col_letter}{row_idx}"].number_format = "0.00%"

            # Red → white (at 0) → green color scale
            ws.conditional_formatting.add(
                data_range,
                ColorScaleRule(
                    start_type="min",
                    start_color="FF4444",   # red = most negative
                    mid_type="num",
                    mid_value=0,
                    mid_color="FFFFFF",     # white = zero
                    end_type="max",
                    end_color="00AA44",     # green = most positive
                ),
            )

        # Auto-fit column width (capped at 20)
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, max_row + 1)),
            default=0,
        )
        ws.column_dimensions[col_letter].width = min(max_len + 2, 20)

    # --- Freeze header ---
    ws.freeze_panes = "A2"

    wb.save(output_path)
    log.info(
        "Returns dashboard saved to %s (%d tickers × %d windows per price type)",
        output_path,
        len(df),
        len(LOOKBACK_WINDOWS),
    )


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------


def main() -> None:
    data_dict = load_results(RESULTS_PATH)

    log.info(
        "Building returns dashboard for %d tickers, windows=%s",
        len(data_dict),
        LOOKBACK_WINDOWS,
    )
    dashboard = build_returns_dashboard(data_dict, windows=LOOKBACK_WINDOWS)

    nan_count = dashboard[[c for c in dashboard.columns if c.startswith("ret_")]].isna().sum().sum()
    if nan_count > 0:
        log.warning(
            "%d NaN return values (tickers with insufficient history for some windows).",
            nan_count,
        )

    log.info("Top 5 tickers by 30-day close return:")
    top5 = dashboard.nlargest(5, "ret_close_30d")[["ticker", "last_date", "close", "ret_close_30d"]]
    log.info("\n%s", top5.to_string(index=False))

    save_returns_dashboard(dashboard, OUTPUT_PATH)


if __name__ == "__main__":
    main()
