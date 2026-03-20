"""
Trending ticker dashboard.

Ranks Italian equity tickers by relative trend score:
  score_vs_bm:     sum of signal values on rclose (ticker / FTSEMIB.MI)
  score_vs_sector: sum of signal values on ticker / sector_peer_avg_close (exclude-self)
  score_total:     score_vs_bm + score_vs_sector

Score range per component: [-N, +N] where N = number of detected signal columns.
A ticker with score_total = 2*N is fully long on all signals vs both benchmark
and its sector peers — maximum conviction.

Usage:
    python trending_dashboard.py

Output:
    data/results/it/trending_dashboard.xlsx
"""

import logging
import sys
from pathlib import Path

import pandas as pd

sys.stdout.reconfigure(encoding="utf-8")
sys.stderr.reconfigure(encoding="utf-8")

from pipeline import load_config, build_search_spaces
from trend_scorer import (
    compute_score_vs_bm,
    compute_score_history,
    compute_score_vs_sector,
    compute_supporting_metrics,
    build_dashboard,
    detect_signal_columns,
)

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------

CONFIG_PATH = Path("config.json")
RESULTS_PATH = Path("data/results/it/analysis_results.parquet")
OHLC_PATH = Path("data/ohlc/historical/it/ohlc_data.parquet")
SECTORS_PATH = Path("data/ticker/it/sectors.xlsx")
OUTPUT_PATH = Path("data/results/it/trending_dashboard.xlsx")
IMAGE_OUTPUT_PATH = Path("data/results/it/trending_heatmap.png")

# Number of trailing trading dates shown in the score heatmap sheet and image.
HEATMAP_DAYS = 60

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)
log = logging.getLogger(__name__)
logging.getLogger("algoshort").setLevel(logging.WARNING)


# ---------------------------------------------------------------------------
# Data loading
# ---------------------------------------------------------------------------


def load_results(path: Path) -> dict[str, pd.DataFrame]:
    """
    Load analysis_results.parquet and partition by symbol.

    Returns:
        symbol -> DataFrame dict. Each DataFrame contains signal columns,
        relative prices (rclose, ropen, rhigh, rlow), and stop_loss columns.

    Raises:
        FileNotFoundError: If parquet file does not exist.
    """
    if not path.exists():
        raise FileNotFoundError(
            f"Results file not found: {path}\n"
            "Run analyze_stock.py first to generate analysis_results.parquet."
        )
    log.info("Loading analysis results from %s", path)
    df = pd.read_parquet(path)
    if "date" in df.columns:
        df["date"] = pd.to_datetime(df["date"])
    data_dict = {symbol: grp.reset_index(drop=True) for symbol, grp in df.groupby("symbol")}
    log.info("Loaded %d symbols", len(data_dict))
    return data_dict


def load_sectors(path: Path) -> pd.DataFrame:
    """
    Load ticker -> sector mapping from sectors.xlsx.

    Expected columns: ticker, name, sector.
    Deduplicates on ticker (first occurrence wins).

    Raises:
        FileNotFoundError: If the file does not exist.
        ValueError: If required columns are missing.
    """
    if not path.exists():
        raise FileNotFoundError(f"Sectors file not found: {path}")
    df = pd.read_excel(path)
    required = {"ticker", "sector"}
    missing = required - set(df.columns)
    if missing:
        raise ValueError(f"sectors.xlsx is missing required columns: {missing}")
    return df[["ticker", "name", "sector"]].drop_duplicates("ticker").reset_index(drop=True)


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------


def save_dashboard(df: pd.DataFrame, output_path: Path) -> None:
    """
    Write the ranked dashboard to Excel with conditional formatting.

    Formatting:
      - Header row: dark-blue background, white bold text
      - score_total column: red-white-green color scale (min → 0 → max)
      - Header row frozen
      - Column widths auto-fitted (capped at 40)

    Args:
        df: Ranked dashboard DataFrame from build_dashboard().
        output_path: Destination .xlsx path. Parent directories are created.
    """
    from openpyxl import load_workbook
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import Alignment, Font, PatternFill

    output_path.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(output_path, index=False, sheet_name="Trending")

    wb = load_workbook(output_path)
    ws = wb["Trending"]

    # Header row formatting
    header_fill = PatternFill(fill_type="solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # Locate score_total column for color scale
    score_total_col_letter = None
    for col_idx, cell in enumerate(ws[1], start=1):
        if cell.value == "score_total":
            from openpyxl.utils import get_column_letter
            score_total_col_letter = get_column_letter(col_idx)
            break

    if score_total_col_letter and ws.max_row > 1:
        score_range = f"{score_total_col_letter}2:{score_total_col_letter}{ws.max_row}"
        ws.conditional_formatting.add(
            score_range,
            ColorScaleRule(
                start_type="min",
                start_color="FF4444",  # red = most negative
                mid_type="num",
                mid_value=0,
                mid_color="FFFFFF",  # white = neutral
                end_type="max",
                end_color="00AA44",  # green = most positive
            ),
        )

    # Freeze header row
    ws.freeze_panes = "A2"

    # Auto-fit column widths
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

    wb.save(output_path)
    log.info("Dashboard saved to %s (%d tickers)", output_path, len(df))


# ---------------------------------------------------------------------------
# Heatmap sheet
# ---------------------------------------------------------------------------


def save_heatmap_sheet(
    score_history: pd.DataFrame,
    ranked_tickers: list[str],
    workbook_path: Path,
) -> None:
    """
    Append a "Heatmap" sheet to an existing workbook.

    Layout:
      - Row 0: header — "ticker" then date columns (formatted as YYYY-MM-DD)
      - Rows 1+: one row per ticker, ordered by ``ranked_tickers`` (dashboard rank order)
      - Values: integer score_vs_bm for each (ticker, date) cell
      - Conditional formatting: red-white-green color scale, min=-N, mid=0, max=N
        where N = number of signal columns (max possible score per date)

    Args:
        score_history: Output of ``compute_score_history`` — wide DataFrame
            with ticker index and date columns.
        ranked_tickers: Tickers in dashboard rank order (top score first).
            Tickers not present in score_history are silently skipped.
        workbook_path: Path to the existing .xlsx file (modified in-place).
    """
    from openpyxl import load_workbook
    from openpyxl.formatting.rule import ColorScaleRule
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = load_workbook(workbook_path)

    # Remove stale heatmap sheet if re-running.
    if "Heatmap" in wb.sheetnames:
        del wb["Heatmap"]
    ws = wb.create_sheet("Heatmap")

    # Filter and reorder tickers to dashboard rank order.
    available = score_history.index
    ordered_tickers = [t for t in ranked_tickers if t in available]
    dates = list(score_history.columns)

    # --- Header row ---
    header_fill = PatternFill(fill_type="solid", fgColor="1F3864")
    header_font = Font(bold=True, color="FFFFFF")

    ws.cell(row=1, column=1, value="ticker").fill = header_fill
    ws.cell(row=1, column=1).font = header_font
    ws.cell(row=1, column=1).alignment = Alignment(horizontal="center")

    for col_idx, dt in enumerate(dates, start=2):
        cell = ws.cell(row=1, column=col_idx, value=dt.strftime("%Y-%m-%d") if hasattr(dt, "strftime") else str(dt))
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    # --- Data rows ---
    for row_idx, ticker in enumerate(ordered_tickers, start=2):
        ws.cell(row=row_idx, column=1, value=ticker)
        for col_idx, dt in enumerate(dates, start=2):
            val = score_history.loc[ticker, dt]
            ws.cell(row=row_idx, column=col_idx, value=int(val) if pd.notna(val) else None)

    # --- Color scale: red → white (0) → green, applied to all data cells ---
    last_row = ws.max_row
    last_col = ws.max_column
    if last_row > 1 and last_col > 1:
        data_range = (
            f"B2:{get_column_letter(last_col)}{last_row}"
        )
        n_signals = len(dates) and score_history.values.max()  # actual max in data
        ws.conditional_formatting.add(
            data_range,
            ColorScaleRule(
                start_type="min",
                start_color="FF4444",   # red = most negative
                mid_type="num",
                mid_value=0,
                mid_color="FFFFFF",     # white = neutral
                end_type="max",
                end_color="00AA44",     # green = most positive
            ),
        )

    # --- Freeze ticker column and header row ---
    ws.freeze_panes = "B2"

    # Ticker column width
    ws.column_dimensions["A"].width = 14

    # Narrow date columns
    for col_idx in range(2, last_col + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 12

    wb.save(workbook_path)
    log.info(
        "Heatmap sheet written: %d tickers × %d dates",
        len(ordered_tickers),
        len(dates),
    )


# ---------------------------------------------------------------------------
# Heatmap image
# ---------------------------------------------------------------------------


def save_heatmap_image(
    score_history: pd.DataFrame,
    ranked_tickers: list[str],
    output_path: Path,
    step: int = 7,
) -> None:
    """
    Save the score heatmap as a PNG image, sampling one date every ``step``
    trading days.

    Layout:
      - Rows: tickers in dashboard rank order (rank 1 at the top)
      - Columns: every ``step``-th trading date (most recent dates on the right)
      - Cell colour: red (most negative) → white (0) → green (most positive),
        anchored at the actual min/max of the data with 0 as the midpoint
      - Colour bar on the right labelled "score_vs_bm"

    Args:
        score_history: Wide DataFrame (ticker index × date columns) from
            ``compute_score_history``.
        ranked_tickers: Tickers in dashboard rank order (rank 1 first).
        output_path: Destination .png path. Parent directories are created.
        step: Sample one date every this many trading days (default 7 ≈ weekly).
    """
    import matplotlib.pyplot as plt
    import matplotlib.colors as mcolors
    import numpy as np

    # --- Prepare data ---
    available = set(score_history.index)
    ordered = [t for t in ranked_tickers if t in available]

    # Sample every `step`-th column; columns are already in ascending date order.
    sampled_dates = score_history.columns[::step]
    data = score_history.loc[ordered, sampled_dates]

    n_tickers, n_dates = data.shape
    values = data.values.astype(float)

    # --- Colormap: red → white → green, centred at 0 ---
    cmap = mcolors.LinearSegmentedColormap.from_list(
        "rwg", ["#FF4444", "#FFFFFF", "#00AA44"]
    )
    # Anchor midpoint at 0 even if data is asymmetric.
    abs_max = max(abs(float(np.nanmin(values))), abs(float(np.nanmax(values))), 1)
    norm = mcolors.TwoSlopeNorm(vmin=-abs_max, vcenter=0, vmax=abs_max)

    # --- Figure size: scale with data dimensions ---
    cell_h = 0.18   # inches per ticker row
    cell_w = 0.55   # inches per date column
    fig_h = max(8, n_tickers * cell_h + 2)
    fig_w = max(10, n_dates * cell_w + 3)

    fig, ax = plt.subplots(figsize=(fig_w, fig_h))

    im = ax.imshow(values, aspect="auto", cmap=cmap, norm=norm, interpolation="nearest")

    # --- Axes labels ---
    ax.set_xticks(range(n_dates))
    ax.set_xticklabels(
        [d.strftime("%Y-%m-%d") if hasattr(d, "strftime") else str(d) for d in sampled_dates],
        rotation=45,
        ha="right",
        fontsize=7,
    )
    ax.set_yticks(range(n_tickers))
    ax.set_yticklabels(ordered, fontsize=6.5)

    ax.set_title("Score vs Benchmark — 1 sample per 7 trading days", fontsize=11, pad=10)
    ax.set_xlabel("Date", fontsize=9)
    ax.set_ylabel("Ticker (ranked by current score)", fontsize=9)

    # --- Colour bar ---
    cbar = fig.colorbar(im, ax=ax, fraction=0.02, pad=0.01)
    cbar.set_label("score_vs_bm", fontsize=8)

    fig.tight_layout()
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig.savefig(output_path, dpi=150, bbox_inches="tight")
    plt.close(fig)
    log.info("Heatmap image saved to %s (%d tickers × %d dates)", output_path, n_tickers, n_dates)


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------


def main() -> None:
    cfg = load_config(CONFIG_PATH)
    search_spaces = build_search_spaces(cfg)
    lookback: int = cfg["returns"]["fast"]

    # --- Load data ---
    data_dict = load_results(RESULTS_PATH)
    sectors_df = load_sectors(SECTORS_PATH)

    ticker_to_sector: dict[str, str] = dict(
        zip(sectors_df["ticker"], sectors_df["sector"])
    )

    # Restrict to tickers present in BOTH results and sector mapping.
    valid_tickers = set(data_dict.keys()) & set(ticker_to_sector.keys())
    unmatched = set(data_dict.keys()) - valid_tickers
    if unmatched:
        log.warning(
            "%d tickers in results have no sector mapping and will be excluded: %s",
            len(unmatched),
            sorted(unmatched)[:10],  # show first 10 to keep log readable
        )
    data_dict = {t: data_dict[t] for t in valid_tickers}
    log.info("Processing %d tickers with sector mapping", len(data_dict))

    # --- Detect signals ---
    first_df = next(iter(data_dict.values()))
    signal_cols = detect_signal_columns(first_df)
    if not signal_cols:
        raise RuntimeError(
            "No signal columns detected in analysis_results.parquet. "
            "Each signal must have a companion <signal>_stop_loss column. "
            "Re-run analyze_stock.py to regenerate results."
        )
    log.info("Detected %d signal columns: %s", len(signal_cols), signal_cols)

    # --- Score vs benchmark (fast: reads last row of existing columns) ---
    log.info("Computing score_vs_bm")
    score_bm = compute_score_vs_bm(data_dict, signal_cols)

    # --- Score vs sector (slow: runs generate_signals on sector-relative prices) ---
    log.info("Loading raw OHLC for sector peer average computation")
    ohlc_df = pd.read_parquet(OHLC_PATH)

    log.info(
        "Computing score_vs_sector for %d tickers across %d sectors "
        "(runs signal generation per ticker — may take several minutes)",
        len(valid_tickers),
        len(set(ticker_to_sector[t] for t in valid_tickers)),
    )
    score_sector = compute_score_vs_sector(
        ohlc_df=ohlc_df,
        ticker_to_sector={t: ticker_to_sector[t] for t in valid_tickers},
        search_spaces=search_spaces,
    )

    # --- Supporting metrics ---
    log.info("Computing supporting metrics (lookback=%d bars)", lookback)
    metrics = compute_supporting_metrics(data_dict, signal_cols, lookback)

    # --- Assemble and save ---
    log.info("Assembling dashboard")
    dashboard = build_dashboard(score_bm, score_sector, metrics, sectors_df)

    log.info("Top 10 tickers by score_total:")
    print(dashboard[["rank", "ticker", "name", "sector", "score_total", "score_vs_bm", "score_vs_sector"]].head(10).to_string(index=False))

    save_dashboard(dashboard, OUTPUT_PATH)

    log.info("Computing score history for heatmap (last %d trading days)", HEATMAP_DAYS)
    score_history = compute_score_history(data_dict, signal_cols, n_days=HEATMAP_DAYS)
    save_heatmap_sheet(
        score_history=score_history,
        ranked_tickers=dashboard["ticker"].tolist(),
        workbook_path=OUTPUT_PATH,
    )

    save_heatmap_image(
        score_history=score_history,
        ranked_tickers=dashboard["ticker"].tolist(),
        output_path=IMAGE_OUTPUT_PATH,
        step=7,
    )


if __name__ == "__main__":
    main()
